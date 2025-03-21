from flask import Flask, request, render_template, jsonify, send_file
import requests
import fitz
import re
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import PatternFill
from io import BytesIO
from PIL import Image 

app = Flask(__name__)

PDF_MIME_TYPE = 'application/pdf'
PDF_EXTENSION = '.pdf'
PDF_ENCRYPTED_MESSAGE = "O PDF está protegido por senha. Por favor, insira a senha."
PDF_INCORRECT_PASSWORD_MESSAGE = "Senha incorreta. Não foi possível abrir o PDF."
PDF_NO_TEXT_EXTRACTED_MESSAGE = "Nenhum texto foi extraído do PDF."
PDF_PATTERN_NOT_FOUND_MESSAGE = "Padrão de dados não encontrado no PDF."

def get_location_from_cep(cep):
    cep = cep.replace("-", "")

    # Tenta buscar na API ViaCEP primeiro
    viacep_data = get_location_from_viacep(cep)
    if viacep_data and viacep_data.get("cidade") != "Cidade não encontrada":
        return viacep_data

    # Se ViaCEP falhar, tenta a API Nominatim
    nominatim_data = get_location_from_nominatim(cep)
    if nominatim_data and nominatim_data.get("latitude") != "Não encontrado":
        return nominatim_data

    # Se ambas falharem, retorna um erro
    return {"cidade": "Erro", "latitude": "Erro", "longitude": "Erro"}

def get_location_from_viacep(cep):
    """Busca cidade e estado usando a API ViaCEP."""
    try:
        cep = cep.replace("-", "") 

        # Busca cidade pelo CEP na API ViaCEP
        viacep_response = requests.get(f"https://viacep.com.br/ws/{cep}/json/")
        if viacep_response.status_code == 200:
            viacep_data = viacep_response.json()
            cidade = viacep_data.get("localidade", "Cidade não encontrada")
        else:
            cidade = "Erro ao buscar cidade"

        # Busca latitude e longitude na API Nominatim (OpenStreetMap)
        nominatim_url = f"https://nominatim.openstreetmap.org/search?postalcode={cep}&country=Brazil&format=json"
        nominatim_response = requests.get(nominatim_url, headers={"User-Agent": "Mozilla/5.0"})
        
        if nominatim_response.status_code == 200 and nominatim_response.json():
            location_data = nominatim_response.json()[0]
            latitude = location_data["lat"]
            longitude = location_data["lon"]
        else:
            nominatim_url = f"https://nominatim.openstreetmap.org/search?city={cidade}&country=Brazil&format=json"
            nominatim_response = requests.get(nominatim_url, headers={"User-Agent": "Mozilla/5.0"})
            if nominatim_response.status_code == 200 and nominatim_response.json():
                location_data = nominatim_response.json()[0]
                latitude = location_data["lat"]
                longitude = location_data["lon"]
            else:
                latitude = "Não encontrado"
                longitude = "Não encontrado"

        return {"cidade": cidade, "latitude": latitude, "longitude": longitude}

    except Exception as e:
        return {"cidade": "Erro", "latitude": "Erro", "longitude": "Erro"}

def get_location_from_nominatim(cep):
    """Busca latitude e longitude usando a API Nominatim (OpenStreetMap)."""
    try:
        nominatim_url = f"https://nominatim.openstreetmap.org/search?postalcode={cep}&country=Brazil&format=json"
        response = requests.get(nominatim_url, headers={"User-Agent": "Mozilla/5.0"})
        
        if response.status_code == 200 and response.json():
            location_data = response.json()[0]
            display_name = location_data.get("display_name", "")

            # Usa regex para extrair o valor entre o CEP e "Região Geográfica"
            match = re.search(r"^\d{5}-?\d{3},\s*([^,]+?)(?=,\s*Região Geográfica|$)", display_name)
            if match:
                cidade = match.group(1).strip()
            else:
                cidade = "Cidade não encontrada"

            latitude = location_data["lat"]
            longitude = location_data["lon"]
            return {"cidade": cidade, "latitude": latitude, "longitude": longitude}
        else:
            return None
    except Exception as e:
        print(f"Erro ao buscar no Nominatim: {e}")
        return None

def validate_file(file):
    if not file:
        return False, "No file part"
    if file.filename == '':
        return False, "No selected file"
    if not file.filename.endswith(PDF_EXTENSION) or file.mimetype != PDF_MIME_TYPE:
        return False, "Invalid file type. Only PDF files are allowed."
    return True, ""

def extract_text_from_pdf(file, password=None):
    try:
        pdf_document = fitz.open(stream=file.read(), filetype="pdf")

        if pdf_document.is_encrypted:
            if not password:
                return None, PDF_ENCRYPTED_MESSAGE
            if not pdf_document.authenticate(password):
                return None, PDF_INCORRECT_PASSWORD_MESSAGE

        text = ""
        for page_num in range(len(pdf_document)):
            page = pdf_document.load_page(page_num)
            page_text = page.get_text("text")
            if page_text:
                text += page_text.strip() + "\n"

        if not text.strip():
            return None, PDF_NO_TEXT_EXTRACTED_MESSAGE

        return text, None
    except Exception as e:
        return None, f"Erro ao processar o PDF: {str(e)}"

def parse_extracted_text(text):
    pattern = r"([A-Z\s]+)\n([A-Z\s0-9]+)\n([A-Z\s]+)\n(?:CEP:\s*)?(\d{5}-?\d{3}|\d{8})\n(?:CPF/CNPJ:\s*)?([\d\.-]+)"
    match = re.search(pattern, text)

    if not match:
        pattern = (
            r"\s*([A-Z\s]+)\n"
            r"\s*([A-Za-z\s0-9]+)\n"
            r"\s*([A-Z\s]+)\n"
            r"\s*CEP:\s*(\d{5}-?\d{3}|\d{8})\n"
            r"\s*CPF/CNPJ:\s*((?:\d{3}\.\d{3}\.\d{3}-\d{2})|(?:\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})|\d{11}|\d{14})"
        )
        match = re.search(pattern, text)

    if not match:
        return None, PDF_PATTERN_NOT_FOUND_MESSAGE

    nome = match.group(1).strip()
    endereco = match.group(2).strip()
    numero_residencia_match = re.search(r"(\d{3,5}\s*(?:[A-Za-z]+\s*\d*\s*[A-Za-z]*\s*\d*)*)", endereco)
    numero_residencia = numero_residencia_match.group(1).strip() if numero_residencia_match else ""
    endereco = endereco.replace(numero_residencia, "").strip()

    bairro_cidade_estado = match.group(3).strip()
    cep = match.group(4).strip()
    cpf_cnpj = match.group(5).strip()

    localizacao = get_location_from_cep(cep)
    cidade = localizacao["cidade"].upper()
    estado = bairro_cidade_estado.split()[-1]
    bairro_estado = bairro_cidade_estado.replace(cidade, "").replace(estado, "").strip()

    tipo_fase = "MONOFÁSICO" if "MONOFÁSICO" in text else "TRIFÁSICO" if "TRIFÁSICO" in text else "BIFÁSICO"
    tensao_atendimento = "220V" if tipo_fase in ["MONOFÁSICO", "BIFÁSICO"] else "380V"

    values_to_match = ["A1", "A2", "A3", "A3a", "A4", "AS", "B1", "B2", "B3", "B4"]

    pattern = r"\b(" + "|".join(re.escape(value) for value in values_to_match) + r")\b"

    match = re.search(pattern, text)
    # Extrai o valor correspondente, se encontrado
    classificacao_unidade = match.group(1).strip() if match else ""
    print(classificacao_unidade)

    numero_cliente_pattern = r"(\d{7,10})\n\d{2}/\d{4}"
    numero_cliente_match = re.search(numero_cliente_pattern, text)
    numero_cliente = numero_cliente_match.group(1).strip() if numero_cliente_match else ""

    extracted_data = {
        "nome": nome,
        "endereco": endereco,
        "numero_residencia": numero_residencia,
        "bairro": bairro_estado,
        "cidade": cidade,
        "latitude": abs(float(localizacao["latitude"])),
        "longitude": abs(float(localizacao["longitude"])),
        "estado": estado,
        "cep": cep,
        "cpf_cnpj": cpf_cnpj,
        "tipo_fase": tipo_fase,
        "tensao_atendimento": tensao_atendimento,
        "classificacao_unidade": classificacao_unidade,
        "numero_cliente": numero_cliente,
    }

    return extracted_data, None

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/extract-text', methods=['POST'])
def extract_text():
    if 'file' not in request.files:
        return jsonify({"error": "No file part"}), 400

    file = request.files['file']
    password = request.form.get('password')

    is_valid, error_message = validate_file(file)
    if not is_valid:
        return jsonify({"error": error_message}), 400

    text, error_message = extract_text_from_pdf(file, password)
    if error_message:
        return jsonify({"error": error_message}), 400

    extracted_data, error_message = parse_extracted_text(text)
    print(extracted_data)
    if error_message:
        return jsonify({"error": error_message}), 400

    return jsonify(extracted_data)

@app.route('/export-to-excel', methods=['POST'])
def export_to_excel():
    nome = request.form.get('nome')
    endereco = request.form.get('endereco')
    bairro = request.form.get('bairro')
    numero_residencia = request.form.get('numero_residencia')
    cidade = request.form.get('cidade')
    cep = request.form.get('cep')
    cpf_cnpj = request.form.get('cpf_cnpj')
    tipo_fase = request.form.get('tipo_fase')
    residencial_conv = request.form.get('classificacao_unidade')
    numero_cliente = request.form.get('numero_cliente')
    latitude = request.form.get('latitude')
    longitude = request.form.get('longitude')
    potencia_instalada = request.form.get('potencia_instalada')
    tensao_atendimento = request.form.get('tensao_atendimento')
    email = request.form.get('email')

    try:
        wb = load_workbook("teste.xlsx")
        ws = wb.active
    except FileNotFoundError:
        return jsonify({"error": "Arquivo teste.xlsx não encontrado."}), 404

    ws['B9'] = f"Código da UC: {numero_cliente}"
    ws['B10'] = f"Titular da UC: {nome}"
    ws['B11'] = f"Rua/Av.: {endereco}"
    ws['B12'] = f"Bairro: {bairro}"
    ws['R11'] = f"Nº: {numero_residencia}"
    ws['J12'] = f"Cidade: {cidade}"
    ws['B13'] = f"Email: {email}"
    ws['B15'] = f"CNPJ/CPF: {cpf_cnpj}"
    ws['Y9'] = residencial_conv
    ws['R19'] = f"Tensão de atendimento (V): {tensao_atendimento}"
    ws['Y11'] = f"CEP: {cep}"
    ws['L18'] = f"(-){latitude}°"
    ws['T18'] = f"(-){longitude}°"
    ws['B19'] = f"Potência instalada (kW): {potencia_instalada}KW"
    ws['B25'] = f"Potência instalada de geração (kW): {potencia_instalada}kW"
    ws['B48'] = f"Nome/Procurador Legal: {nome}"
    ws['B50'] = f"E-mail: {email}"

    fill_black = PatternFill(start_color="000000", end_color="000000", fill_type="solid") 

    if 'B' in residencial_conv:
        ws['J9'].fill = fill_black
    elif 'A' in residencial_conv:
        ws['R9'].fill = fill_black

    if 'MONOFÁSICO' == tipo_fase:
        ws['J20'].fill = fill_black
    elif 'BIFÁSICO' == tipo_fase:
        ws['R20'].fill = fill_black
    else:
        ws['Y20'].fill = fill_black
    
    ws['J22'].fill = fill_black
    ws['AE27'].fill = fill_black

    try:
        img = Image.open("enel.png")  

        width, height = img.size
        new_width = int(width * 0.6)
        new_height = int(height * 0.6)
        img_resized = img.resize((new_width, new_height))

        img_resized_path = "enel_resized.png"
        img_resized.save(img_resized_path)

        img_excel = ExcelImage(img_resized_path)
        ws.add_image(img_excel, 'B2')
    except FileNotFoundError:
        return jsonify({"error": "Arquivo enel.png não encontrado."}), 404

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return send_file(
        excel_file,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="dados_extraidos.xlsx"
    )

if __name__ == '__main__':
    app.run(debug=True)